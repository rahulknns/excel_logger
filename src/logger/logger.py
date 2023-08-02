import openpyxl as xl
import datetime as dt
import os

class Logger:
    def __init__(self,path = "~",prefix = "log",name = None,overwrite=False,meta_data = None,headers:list= None):
        """!
            @brief: Constructor for the Logger class.
            @description: This function is used to initialize the Logger class. It creates a new excel file if the file does not exist.
            @param: path: path where the file is to be saved. should be a string.
            @param: prefix: prefix of the file name. should be a string.
            @param: name: name of the file. should be a string.
            @param: overwrite: if True then the file will be overwritten.
            @param: meta_data: meta data to be saved in the file before the data.
            @param: headers: headers of the data to be saved in the file.
            @return: None
        """
        self.path = os.path.expanduser(path)
        self.prefix = prefix
        self.name = name
        self.counter = 0
        if self.name is None:
            self.name = dt.datetime.now().strftime("%Y-%m-%d")    
        if not os.path.exists(self.path):
            os.makedirs(self.path)
        self.file_name = self.prefix + "_" + self.name + ".xlsx"
        self.file_path = os.path.join(self.path,self.file_name)
        self.file_path = os.path.expanduser(self.file_path)
        if  os.path.exists(self.file_path):
            try:
                if overwrite:
                    self.wb = xl.Workbook()
                else:
                    self.wb = xl.load_workbook(self.file_path)
            except:
                self.wb = xl.Workbook()
                self.file_name = self.prefix + "_" + self.name + "_" + dt.datetime.now().strftime("%H-%M-%S") + ".xlsx"
                self.file_path = os.path.join(self.path,self.file_name)
                self.file_path = os.path.expanduser(self.file_path)
        else:
            self.wb = xl.Workbook()
        self.ws = self.wb.create_sheet(dt.datetime.now().strftime("%H-%M-%S"))
        if meta_data is not None:
            self.ws.append(list(meta_data.keys()))
            self.ws.append( list(meta_data.values()) )
        self.ws.append([])
        if headers is not None:
            self.ws.append(["SI No" , "Time"] + headers)
        else:
            self.ws.append(["SI No","Time","Data"])
    

    def log(self,data:list):
        """!
            @brief: This function is used to log data in excel file.
            @param: data: data to be logged.
            @return: None
        """
        self.counter += 1
        self.ws.append([self.counter,dt.datetime.now().strftime("%H:%M:%S")] + data)    
    

    def save(self):
        """!
            @brief: This function is used to save the file.
            @param: None
            @return: None
        """
        self.wb.save(self.file_path)

    def __enter__(self):
        return self 
    
    def __exit__(self,exc_type,exc_value,exc_traceback):
        self.save()
    
